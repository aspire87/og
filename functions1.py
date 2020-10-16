import openpyxl as oxl
from openpyxl.utils import column_index_from_string as col_index
from openpyxl.styles import PatternFill
import requests


# import re


# получение данных из столбца(стартовая строка,стобец(букв), имя файла)
# на выходе список
def take_all_data_fr_col_to_list(start_row, start_col, filename):
    start_row = int(start_row)
    col = col_index(start_col.upper())
    wb = oxl.load_workbook(filename)
    sheet = wb.active
    full_addr_list = []
    for full_addr in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=col, max_col=col,
                                     values_only=True):
        for data_split in full_addr:
            full_addr_list.append(data_split)
    return full_addr_list


# заполнение ячеек  вложенным списком построчно
# каждый вложенный список соотвествует строке
# если значение ==' ' или '#' красит ячейку красным
def fill_cells_by_list(start_row, start_col, filename, lst):
    start_row = int(start_row)
    redfill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
    start_col = col_index(start_col.upper())
    wb = oxl.load_workbook(filename)
    sheet = wb.active
    for row in range(0, len(lst)):
        for col in range(0, len(lst[row])):
            if lst[row][col] == ' ' or lst[row][col] == '#':
                sheet.cell(row=start_row + row, column=start_col + col).fill = redfill
            sheet.cell(row=start_row + row, column=start_col + col).value = lst[row][col]
    wb.save(filename)


# заполнение одной строки одноуровневым списком
def fill_cells_by_small_list(start_row, start_col, filename, lst):
    start_col = col_index(start_col.upper())
    wb = oxl.load_workbook(filename)
    sheet = wb.active
    for col in range(0, len(lst)):
        sheet.cell(row=start_row, column=start_col + col).value = lst[col]
    wb.save(filename)


# в случае нескольких соответсвий адреса-предлагает пользователю выбрать вариант
# принимает json ответ, возвращает словарь
def choose_variable(response):
    members = response['response']['GeoObjectCollection']['featureMember']
    req = response['response']['GeoObjectCollection']['metaDataProperty']['GeocoderResponseMetaData']['request']
    print(' ')
    print(f'{req}')
    print('Выберите наиболее подходящий:')
    for member in members:
        print(members.index(member) + 1, member['GeoObject']['metaDataProperty']['GeocoderMetaData']['text'])
    choose = int(input('Введите номер строки: '))
    adress_output = {'city': ' ', 'street': ' ', 'house': ' ', 'coord1': ' ',
                     'coord2': ' ', 'reg_name': ' '}
    try:
        coordinates = response['response']['GeoObjectCollection']['featureMember'][choose - 1]['GeoObject']['Point'][
            'pos']
    except IndexError:
        coordinates = '# #'
    coordinates = coordinates.split()
    adress_output['coord1'], adress_output['coord2'] = coordinates
    try:
        address_full = response['response']['GeoObjectCollection']['featureMember'][choose - 1]['GeoObject'][
            'metaDataProperty']['GeocoderMetaData']['Address']['Components']
        for data in address_full:
            for val in data.values():
                if 'locality' in val:
                    adress_output['city'] = data.get('name')
                elif 'street' in val:
                    adress_output['street'] = data.get('name')
                elif 'house' in val:
                    adress_output['house'] = data.get('name')
        if adress_output['street'] == ' ':
            for data in address_full:
                for val in data.values():
                    if 'district' in val:
                        adress_output['street'] = data.get('name')
        region = response['response']['GeoObjectCollection']['featureMember'][choose - 1]['GeoObject'][
            'metaDataProperty']['GeocoderMetaData']['Address']['Components'][2]
        adress_output['reg_name'] = region.get('name')
    except IndexError:
        return adress_output
    return adress_output


# получение координат,точного адреса по приблизительному адресу
# возвращает  словарь{город: улица: дом: коорд1: коорд2: названиеобл:}
def get_coordinats(addr):
    adress_output = {'city': ' ', 'street': ' ', 'house': ' ', 'coord1': ' ',
                     'coord2': ' ', 'reg_name': ' ', }
    api_key = '80c29bc8-378a-47e9-b6ba-9e755c788100'
    req = "https://geocode-maps.yandex.ru/1.x/?format=json&apikey=" + api_key + "&lang=ru_RU&geocode=" + addr
    res = requests.get(req)
    res = res.json()
    founded = int(res['response']['GeoObjectCollection'][
                      'metaDataProperty']['GeocoderResponseMetaData'].get('found'))
    # тегирование неоднозначного адреса
    if founded > 1:
        return {'bad_result': True, 'res': res}
    elif founded < 1:
        adress_output = {'city': ' ', 'street': ' ', 'house': ' ', 'coord1': ' ',
                         'coord2': ' ', 'reg_name': ' ', 'bad_result': False, 'not_founded': True, 'res': res}
        return adress_output
    else:
        adress_output['bad_result'] = False
        try:
            coordinates = res['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point'].get('pos')
        except IndexError:
            coordinates = '# #'
        adress_output['coord1'], adress_output['coord2'] = coordinates.split()
        try:
            address_full = res['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                'metaDataProperty']['GeocoderMetaData']['Address']['Components']
        except IndexError:
            return adress_output
        for data in address_full:
            for val in data.values():
                if 'locality' in val:
                    adress_output['city'] = data.get('name')
                elif 'street' in val:
                    adress_output['street'] = data.get('name')
                elif 'house' in val:
                    adress_output['house'] = data.get('name')
        if adress_output.get('street') == ' ':
            for data in address_full:
                for val in data.values():
                    if 'district' in val:
                        adress_output['street'] = data.get('name')
        region = res['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
            'metaDataProperty']['GeocoderMetaData']['Address']['Components'][2]
        adress_output['reg_name'] = region.get('name')
    return adress_output


#  убираем слово "улица", заменяем строение на слеши
# принимает словарь, возвращает список
def parse_address(add):
    street_type_list = ['площадь', 'просека', 'аллея', "бульвар", "набережная", "переулок", "проезд", "проспект",
                        "спуск", "тупик", "шоссе", "улица", "тракт", "имени",'микрорайон ']
    city_type_list=['посёлок городского типа ','деревня ','посёлок ','рабочий посёлок ']
    city = add.get('city')
    street = add.get('street')
    house = add.get('house')
    #поиск города
    for city_type in city_type_list:
        if city_type in street:
            city = city.replace(city_type, '').strip()
    add['city'] = city
    # поиск улицы
    for street_type in street_type_list:
        if street_type in street:
            street = street.replace(street_type, '').strip()
    add['street'] = street
    # замена строений на слеши
    if not house.isdigit():
        for letter in house:
            if not letter.isdigit() and not letter.isupper():
                add['house'] = house.replace(letter, '/')
    add_string = add.get('city') + ' ' + add.get('street') + ' ' + add.get('house')
    return add, add_string


# принимает строку, возвращает строку транслитом
# учитывает регистр
def transliteration(line):
    out_line = ''
    alph = {'а': 'a', "б": 'b', "в": 'v', "г": 'g', 'д': 'd', 'е': "e",
            'ё': "e", "ж": "zh", 'з': "z", 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o',
            'п': 'p', 'р': 'r',
            'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'h', 'ц': 'c', 'ч': 'ch', 'ш': 'sh', 'щ': 'sh', 'э': 'e',
            'ю': 'u', 'я': 'ya', 'ы': 'y', ' ': '_', '.': '_', '_': '_', '-': '_', '/': '_'}
    for word in line.split():
        word += '_'
        for el in word:
            if el.isdecimal():
                out_line += el
            elif not el.isalpha():
                el = alph.get(el)
                if el is not None:
                    out_line += el
            else:
                if el.isupper():
                    el = alph.get(el.lower())
                    if el is not None and len(el) == 2:
                        out_line += el[0].upper() + el[1].lower()
                    elif el is not None:
                        out_line += el.upper()
                else:
                    el = alph.get(el.lower())
                    if el is not None:
                        out_line += el
    out_line = out_line.rstrip('_')
    # out_line=re.sub(r'\w_\w_','_',out_line)
    # out_line=re.sub(r'_\w_\w_','_',out_line)
    # out_line=re.sub(r'\b\w_\B','_',out_line)
    # out_line=re.sub(r'__','_',out_line)

    return out_line


# принимает путь к файлу TXT, отдает словарь{регион:01}
def create_regions_dict_from_txt(path):
    reg_dict = {}
    file = open(path)
    data = file.readlines()
    for line in data:
        if line == '\n':
            continue
        else:
            lin = line.rstrip('\n').lower().split(':')
            reg_dict[lin[1]] = lin[0]
    file.close()
    return reg_dict


# поиск региона в словаре
# принимает словарь и строку, отдает строку
def get_reg_num(dictt, line):
    reg_num = ''
    for key in dictt.keys():
        if line.lower() in key.lower():
            reg_num = dictt.get(key)
            return reg_num
        else:
            reg_num = ' '
    return reg_num
# функция вывода не найденных результатов
# def not_founded_output(res):
#  req = res['response']['GeoObjectCollection']['metaDataProperty']['GeocoderResponseMetaData']['request']
# print(f'Внимание для данного запроса '
#      f'{req}'
#     'не нашлось ни одного результата')
