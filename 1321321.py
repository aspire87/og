import openpyxl
import datetime
from openpyxl.utils import column_index_from_string as col_index
from os import path

klient = 'ООО Рога и копыта'
nowdate = datetime.datetime.today().strftime('%d.%m.%Y')
dateout = ''
mrf = 'Центр'
gorod = 'Тула'
city = 'Ефремов'
ulica = 'Лермонтова'
dom = '36'
td = '2'
cms = 'Организован'
zakaz = '20-753952'
domen = 'OOO_Roga_i_kopyta_Lermontova_36'
mac = 'e0:d9:e3:70:47:24'
if str(gorod) == str(city):
    addr = [klient, nowdate, dateout, mrf, gorod, ulica + ', ' + dom, td, cms, zakaz, domen, mac]
else:
    addr = [klient, nowdate, dateout, mrf, gorod, city + ', ' + ulica + ', ' + dom, td, cms, zakaz, domen, mac]
filename = nowdate + '.xlsx'
wb = openpyxl.load_workbook('template.xlsx')
sheet = wb.active
max_row = sheet.max_row + 1
start_col = 'A'
start_col = col_index(start_col)
for col in range(0, len(addr)):
        sheet.cell(row=max_row, column=start_col + col).value = addr[col]
wb.save('template.xlsx')
