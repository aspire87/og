import openpyxl as oxl
from openpyxl.utils import column_index_from_string as col_index
from os import path
from work_with_dic_states import date
# from openpyxl.styles import PatternFill
from datetime import datetime


def chck_file_exist(func):
    def inner_func(self,lst):
        header=['Наименование клиента','Дата подключения',
                'Дата отключения','МРФ','Город','Адрес предоставления услуги',
                'Количество точек доступа','Статус CMS','Номер заказа CMS',
                'Домен','МАС']
        if path.exists(self.filename):
            return func(self,lst)
        else:
            wb = oxl.Workbook()
            sheet = wb.active
            for col in range(0, len(header)):
                sheet.cell(row=1, column=self.start_col+col).value = header[col]
            wb.save(self.filename)
            return func(self,lst)

    return inner_func

class Excel():

    def __init__(self, start_col='A'):
        self.date = date.strftime('%d-%m-%Y')
        self.filename = self.date + '.xlsx'
        # self.wb = oxl.load_workbook(self.filename)
        # self.sheet = wb.active
        # self.max_row = self.sheet.max_row
        self.start_col = col_index(start_col)



    @chck_file_exist
    def write_list(self, lst):
        wb = oxl.load_workbook(self.filename)
        sheet = wb.active
        max_row=sheet.max_row+1
        for col in range(0, len(lst)):
            sheet.cell(row=max_row, column=self.start_col + col).value = lst[col]
        wb.save(self.filename)


lst = ['a', 'f', 'd']
exc = Excel()
exc.write_list(lst)
